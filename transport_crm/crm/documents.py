"""
Модуль генерации документов (путевой лист, счёт, акт) в формате Excel.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Order

def generate_waybill(order: Order) -> HttpResponse:
    """
    Генерирует путевой лист для заказа.

    Args:
        order: Экземпляр заказа.

    Returns:
        HttpResponse с файлом Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Путевой лист"

    ws.merge_cells('A1:F1')
    ws['A1'] = f"ПУТЕВОЙ ЛИСТ №{order.id}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    data = [
        ("Дата", order.requested_date.strftime('%d.%m.%Y') if order.requested_date else ""),
        ("Клиент", order.customer.name),
        ("Адрес подачи", order.pickup_address),
        ("Адрес доставки", order.delivery_address),
        ("Груз", order.cargo_description),
        ("Вес, т", float(order.weight_ton)),
        ("Расстояние, км", float(order.distance_km) if order.distance_km else ""),
    ]
    if hasattr(order, 'assignment') and order.assignment:
        data += [
            ("Водитель", f"{order.assignment.driver.first_name} {order.assignment.driver.last_name}"),
            ("Автомобиль", f"{order.assignment.vehicle.plate_number} {order.assignment.vehicle.brand} {order.assignment.vehicle.model}"),
        ]

    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="waybill_{order.id}.xlsx"'
    return response


def generate_invoice(order: Order) -> HttpResponse:
    """
    Генерирует счёт на оплату для заказа.

    Args:
        order: Экземпляр заказа.

    Returns:
        HttpResponse с файлом Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Счёт"

    ws.merge_cells('A1:D1')
    ws['A1'] = f"СЧЁТ №{order.id}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    data = [
        ("Дата", order.requested_date.strftime('%d.%m.%Y') if order.requested_date else ""),
        ("Заказчик", order.customer.name),
        ("Описание услуги", f"Перевозка груза: {order.cargo_description}"),
        ("Маршрут", f"{order.pickup_address} — {order.delivery_address}"),
        ("Стоимость, руб.", float(order.price) if order.price else ""),
        ("НДС не облагается", ""),
    ]
    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="invoice_{order.id}.xlsx"'
    return response


def generate_act(order: Order) -> HttpResponse:
    """
    Генерирует акт выполненных работ для заказа.

    Args:
        order: Экземпляр заказа.

    Returns:
        HttpResponse с файлом Excel.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Акт"

    ws.merge_cells('A1:D1')
    ws['A1'] = f"АКТ ВЫПОЛНЕННЫХ РАБОТ №{order.id}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    data = [
        ("Дата выполнения", order.requested_date.strftime('%d.%m.%Y') if order.requested_date else ""),
        ("Заказчик", order.customer.name),
        ("Описание", f"Перевозка груза по маршруту {order.pickup_address} — {order.delivery_address}"),
        ("Объём, т", float(order.weight_ton)),
        ("Расстояние, км", float(order.distance_km) if order.distance_km else ""),
        ("Стоимость, руб.", float(order.price) if order.price else ""),
    ]
    for i, (label, value) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="act_{order.id}.xlsx"'
    return response